from openpyxl import Workbook
import random

wb = Workbook()

# grab the active worksheet
ws = wb.active

#количество траифка по позициям
traffic = [1, 0.85, 0.75, 0.65, 0.06, 0.05, 0.04, 0.03, 0.02, 0, 0]

<<<<<<< HEAD
#целевая стоимость конверсии
click_value = 650

=======
>>>>>>> parent of d698773... Работающий оптимизатор v1
#задаем вероятности конверсий и количетсво трафика ключам
ver_conv = [0]*50
amount_traff = [0]*50
for i in range(50): 
    ver_conv[i] = round(random.uniform(0.1, 5)/100, 2)
    amount_traff[i] = random.randint(1, 100)
    
cost_click = [20]*50 #массив ставок по ключам. начальная ставка 20 руб.
<<<<<<< HEAD
<<<<<<< HEAD
=======
>>>>>>> parent of d698773... Работающий оптимизатор v1
base_cost_click = 25 #контрольная фиксированная ставка для всех ключей - 20 руб
=======
base_cost_click = 20 #контрольная фиксированная ставка для всех ключей - 20 руб
>>>>>>> parent of ac182db... бэк
cost_pos = [[0]*9 for i in range(50)] #массив стоимости позиций
all_count_conv = 0
all_money = 0
all_money_fix = 0
u = 0
l = 50
count_conv = [0]*50
all_key_conv = [0]*50
all_key_costs = [0]*50
costs_per_key = [0]*50
full_money = 0
full_conv = 0

full_money_fix = 0
full_conv_fix = 0
full_costs_per_key = [0]*50
full_conv_per_key = [0]*50

<<<<<<< HEAD
<<<<<<< HEAD
full_full_money = 0
full_full_conv = 0
        
full_full_money_fix = 0
full_full_conv_fix = 0

for q in range(10):
    random.seed

    #задаем вероятности конверсий и количетсво трафика ключам
    for i in range(50): 
        ver_conv[i] = round(random.uniform(0.5, 5)/100, 2)
        amount_traff[i] = random.randint(1, 50)

    for x in range (1, 10):
        all_count_conv = 0
        all_count_conv_fix = 0
        money = 0
        money_fix = 0
        traff_period = [0]*50
        for v in range(1, 50):
            #генерируем стоимость позиций для ключа
            cost_pos[v][8] = round(random.uniform(1, 3), 2)
            cost_pos[v][7] = cost_pos[v][8] + round(random.uniform(0.5, 2), 2)
            cost_pos[v][6] = cost_pos[v][7] + round(random.uniform(0.5, 2), 2)
            cost_pos[v][5] = cost_pos[v][6] + round(random.uniform(0.5, 2), 2)
            cost_pos[v][4] = cost_pos[v][5] + round(random.uniform(0.5, 3), 2)
            cost_pos[v][3] = cost_pos[v][4] + round(random.uniform(4, 10), 2)
            cost_pos[v][2] = cost_pos[v][3] + round(random.uniform(2, 10), 2)
            cost_pos[v][1] = cost_pos[v][2] + round(random.uniform(1, 15), 2)
            cost_pos[v][0] = cost_pos[v][1] + round(random.uniform(2, 10), 2)
        
            #записываем в excel цены 
            '''for i in range(1, 10):
                ws.cell(row=v+u, column=i).value = cost_pos[v][i-1]'''

            #ищем максимальную доступную позицию и стоимость клика для оптимизатора
            cpc = [0]*50
            for i in range(9):
                if cost_click[v] == 0:
                    num_pos = 10
                    cpc[v] = 0
                    break
                if cost_pos[v][i] <= cost_click[v]:
                    cpc[v] = cost_pos[v][i]+0.01
                    num_pos = i
                    break

            #максимальная позиция и стоимость клика для фиксированной ставки
            cpc_fix = [0]*50
            for j in range(9):
                if cost_pos[v][j] <= base_cost_click:
                    cpc_fix[v] = cost_pos[v][j]+0.01
                    num_pos_fix = j
                    break
                
            #считаем количество трафика за период для оптимизатора
            traff_period[v] = round(amount_traff[v] * traffic [num_pos], 0)

            #считаем количество трафика за период для фиксированной ставки
<<<<<<< HEAD
            traff_period_fix = int(round(amount_traff[v] * traffic [num_pos_fix], 0))
=======
for x in range (1, 10):
    all_count_conv = 0
    all_count_conv_fix = 0
    money = 0
    money_fix = 0
    traff_period = [0]*50
    for v in range(1, 50):
        #генерируем стоимость позиций для ключа
        cost_pos[v][8] = round(random.uniform(1, 3), 2)
        cost_pos[v][7] = cost_pos[v][8] + round(random.uniform(0.5, 2), 2)
        cost_pos[v][6] = cost_pos[v][7] + round(random.uniform(0.5, 2), 2)
        cost_pos[v][5] = cost_pos[v][6] + round(random.uniform(0.5, 2), 2)
        cost_pos[v][4] = cost_pos[v][5] + round(random.uniform(0.5, 3), 2)
        cost_pos[v][3] = cost_pos[v][4] + round(random.uniform(4, 10), 2)
        cost_pos[v][2] = cost_pos[v][3] + round(random.uniform(2, 10), 2)
        cost_pos[v][1] = cost_pos[v][2] + round(random.uniform(1, 15), 2)
        cost_pos[v][0] = cost_pos[v][1] + round(random.uniform(2, 10), 2)

        #записываем в excel цены 
        for i in range(1, 10):
            ws.cell(row=v+u, column=i).value = cost_pos[v][i-1]

        #ищем максимальную доступную позицию и стоимость клика для оптимизатора
        cpc = [0]*50
        for i in range(9):
            if cost_pos[v][i] <= cost_click[v]:
                cpc[v] = cost_pos[v][i]+0.01
                num_pos = i
                break

        #максимальная позиция и стоимость клика для фиксированной ставки
        cpc_fix = [0]*50
        for j in range(9):
            if cost_pos[v][j] <= base_cost_click:
                cpc_fix[v] = cost_pos[v][j]+0.01
                num_pos_fix = j
                break
            
        #считаем количество трафика за период для оптимизатора
        traff_period[v] = round(amount_traff[v] * traffic [num_pos], 0)
=======
            traff_period_fix = round(amount_traff[v] * traffic [num_pos_fix], 0)
>>>>>>> parent of ac182db... бэк

        #считаем количество трафика за период для фиксированной ставки
        traff_period_fix = round(amount_traff[v] * traffic [num_pos_fix], 0)
>>>>>>> parent of d698773... Работающий оптимизатор v1

        #расходы по ключу оптимизатора
        costs_per_key[v] = traff_period[v] * cpc[v]
        full_costs_per_key[v] += costs_per_key[v]

<<<<<<< HEAD
        #расходы по ключу фикс
        costs_per_key_fix = traff_period_fix * cpc_fix[v]

<<<<<<< HEAD

=======
>>>>>>> parent of ac182db... бэк
            #количество конверсий оптимизатора
            count_conv[v] = round(traff_period[v] * ver_conv[v], 0)
            all_count_conv += count_conv[v]
            full_conv_per_key[v] += count_conv[v]

            #количество конверсий с фикс ставкой
            count_conv_fix = round(traff_period_fix * ver_conv[v], 0)
            all_count_conv_fix += count_conv_fix
=======
        #количество конверсий оптимизатора
        count_conv[v] = round(traff_period[v] * ver_conv[v], 0)
        all_count_conv += count_conv[v]
        full_conv_per_key[v] += count_conv[v]

        #количество конверсий с фикс ставкой
        count_conv_fix = round(traff_period_fix * ver_conv[v], 0)
        all_count_conv_fix += count_conv_fix
>>>>>>> parent of d698773... Работающий оптимизатор v1
=======
for x in range (1, 10):
    all_count_conv = 0
    all_count_conv_fix = 0
    money = 0
    money_fix = 0
    traff_period = [0]*50
    for v in range(1, 50):
        #генерируем стоимость позиций для ключа
        cost_pos[v][8] = round(random.uniform(1, 3), 2)
        cost_pos[v][7] = cost_pos[v][8] + round(random.uniform(0.5, 2), 2)
        cost_pos[v][6] = cost_pos[v][7] + round(random.uniform(0.5, 2), 2)
        cost_pos[v][5] = cost_pos[v][6] + round(random.uniform(0.5, 2), 2)
        cost_pos[v][4] = cost_pos[v][5] + round(random.uniform(0.5, 3), 2)
        cost_pos[v][3] = cost_pos[v][4] + round(random.uniform(4, 10), 2)
        cost_pos[v][2] = cost_pos[v][3] + round(random.uniform(2, 10), 2)
        cost_pos[v][1] = cost_pos[v][2] + round(random.uniform(1, 15), 2)
        cost_pos[v][0] = cost_pos[v][1] + round(random.uniform(2, 10), 2)

        #записываем в excel цены 
        for i in range(1, 10):
            ws.cell(row=v+u, column=i).value = cost_pos[v][i-1]

        #ищем максимальную доступную позицию и стоимость клика для оптимизатора
        cpc = [0]*50
        for i in range(9):
            if cost_pos[v][i] <= cost_click[v]:
                cpc[v] = cost_pos[v][i]+0.01
                num_pos = i
                break

        #максимальная позиция и стоимость клика для фиксированной ставки
        cpc_fix = [0]*50
        for j in range(9):
            if cost_pos[v][j] <= base_cost_click:
                cpc_fix[v] = cost_pos[v][j]+0.01
                num_pos_fix = j
                break
            
        #считаем количество трафика за период для оптимизатора
        traff_period[v] = round(amount_traff[v] * traffic [num_pos], 0)

        #считаем количество трафика за период для фиксированной ставки
        traff_period_fix = round(amount_traff[v] * traffic [num_pos_fix], 0)

        #расходы по ключу оптимизатора
        costs_per_key[v] = traff_period[v] * cpc[v]
        full_costs_per_key[v] += costs_per_key[v]

        #расходы по ключу фикс
        costs_per_key_fix = traff_period_fix * cpc_fix[v]

        #количество конверсий оптимизатора
        count_conv[v] = round(traff_period[v] * ver_conv[v], 0)
        all_count_conv += count_conv[v]
        full_conv_per_key[v] += count_conv[v]

        #количество конверсий с фикс ставкой
        count_conv_fix = round(traff_period_fix * ver_conv[v], 0)
        all_count_conv_fix += count_conv_fix
>>>>>>> parent of d698773... Работающий оптимизатор v1

        conv_cost = [0]*50
        #стоимость конверсии по ключам с опитимизатором
        if count_conv[v] > 0:
            conv_cost[v] = costs_per_key[v] / count_conv[v]

        conv_cost_fix = [0]*50
        #стоимость конверсии по ключам с фискированной ставкой
        if count_conv_fix > 0:
            conv_cost_fix[v] = costs_per_key_fix / count_conv_fix
            
<<<<<<< HEAD
<<<<<<< HEAD
            #фиксированная ставка 
            #ws.cell(row=v+u, column=11).value = cpc_fix[v]
            #ws.cell(row=v+u, column=12).value = num_pos_fix
            #ws.cell(row=v+u, column=13).value = traff_period_fix
            #ws.cell(row=v+u, column=14).value = costs_per_key_fix
            #ws.cell(row=v+u, column=15).value = count_conv_fix
            #ws.cell(row=v+u, column=16).value = conv_cost_fix[v]

            #оптимизированная ставка
            #ws.cell(row=v+u, column=18).value = cpc[v]
            #ws.cell(row=v+u, column=19).value = num_pos
            #ws.cell(row=v+u, column=20).value = traff_period[v]
            #ws.cell(row=v+u, column=21).value = costs_per_key[v]
            #ws.cell(row=v+u, column=22).value = count_conv[v]
            #ws.cell(row=v+u, column=23).value = conv_cost[v]
            #ws.cell(row=v+u, column=24).value = cost_click[v]
            #ws.cell(row=v+u, column=28).value = full_costs_per_key[v]
            #ws.cell(row=v+u, column=29).value = full_conv_per_key[v]
            #if full_conv_per_key[v] > 0:
            #    ws.cell(row=v+u, column=30).value = full_costs_per_key[v] / full_conv_per_key[v]
<<<<<<< HEAD
=======

>>>>>>> parent of d5775dc... бэкап

        #считаем расходы и конверсии у фиксированной ставки
        if all_count_conv != 0:
            cost_conv_period_fix = money_fix / all_count_conv
        else:
            cost_conv_period_fix = money_fix
        all_money_fix += money_fix

        #считаем расходы и конверсии у оптимизатора
        cost_conv_period = money / all_count_conv
        all_money += money

        '''#ввывод итоговых цифр за период
        ws.cell(row=l, column=14).value = money_fix
        ws.cell(row=l, column=21).value = money
        ws.cell(row=l, column=15).value = all_count_conv_fix
        ws.cell(row=l, column=22).value = all_count_conv
        ws.cell(row=l, column=16).value = cost_conv_period_fix
        ws.cell(row=l, column=23).value = cost_conv_period'''

        U = 0
        full_money += money
        full_conv += all_count_conv
=======
        #расходы за период
        money += costs_per_key[v]
        money_fix += costs_per_key_fix
>>>>>>> parent of d698773... Работающий оптимизатор v1
        
        #фиксированная ставка 
        ws.cell(row=v+u, column=11).value = cpc_fix[v]
        ws.cell(row=v+u, column=12).value = num_pos_fix
        ws.cell(row=v+u, column=13).value = traff_period_fix
        ws.cell(row=v+u, column=14).value = costs_per_key_fix
        ws.cell(row=v+u, column=15).value = count_conv_fix
        ws.cell(row=v+u, column=16).value = conv_cost_fix[v]

        #оптимизированная ставка
        ws.cell(row=v+u, column=18).value = cpc[v]
        ws.cell(row=v+u, column=19).value = num_pos
        ws.cell(row=v+u, column=20).value = traff_period[v]
        ws.cell(row=v+u, column=21).value = costs_per_key[v]
        ws.cell(row=v+u, column=22).value = count_conv[v]
        ws.cell(row=v+u, column=23).value = conv_cost[v]
        ws.cell(row=v+u, column=24).value = cost_click[v]
        ws.cell(row=v+u, column=28).value = full_costs_per_key[v]
        ws.cell(row=v+u, column=29).value = full_conv_per_key[v]
        if full_conv_per_key[v] > 0:
            ws.cell(row=v+u, column=30).value = full_costs_per_key[v] / full_conv_per_key[v]


    #считаем расходы и конверсии у фиксированной ставки
    cost_conv_period_fix = money_fix / all_count_conv
    all_money_fix += money_fix

    #считаем расходы и конверсии у оптимизатора
    cost_conv_period = money / all_count_conv
    all_money += money

    #ввывод итоговых цифр за период
    ws.cell(row=l, column=14).value = money_fix
    ws.cell(row=l, column=21).value = money
    ws.cell(row=l, column=15).value = all_count_conv_fix
    ws.cell(row=l, column=22).value = all_count_conv
    ws.cell(row=l, column=16).value = cost_conv_period_fix
    ws.cell(row=l, column=23).value = cost_conv_period

    U = 0
    full_money += money
    full_conv += all_count_conv
    
    full_money_fix += money_fix
    full_conv_fix += all_count_conv_fix
    #выставляем ставки
    for i in range(1, 50):
        if full_conv_per_key[i] > 0:
            if full_costs_per_key[i] / full_conv_per_key[i] < click_value:
                cost_click[i] = cost_click[i] + cost_click[i]*0.2
            else:
<<<<<<< HEAD
                if full_costs_per_key[i] > 1700:
                    cost_click[i] = 0

        u += 52
        l += 52           

        '''
            U = traff_period[i] * ver_conv[i]
    
            if U * cpc[i] < click_value:
                cost_click[i] = cost_click[i] + cost_click[i]*0.1
            else:
                cost_click[i] = cost_click[i] - cost_click[i]*0.1'''
               
=======
        #расходы за период
        money += costs_per_key[v]
        money_fix += costs_per_key_fix
        
        #фиксированная ставка 
        ws.cell(row=v+u, column=11).value = cpc_fix[v]
        ws.cell(row=v+u, column=12).value = num_pos_fix
        ws.cell(row=v+u, column=13).value = traff_period_fix
        ws.cell(row=v+u, column=14).value = costs_per_key_fix
        ws.cell(row=v+u, column=15).value = count_conv_fix
        ws.cell(row=v+u, column=16).value = conv_cost_fix[v]

        #оптимизированная ставка
        ws.cell(row=v+u, column=18).value = cpc[v]
        ws.cell(row=v+u, column=19).value = num_pos
        ws.cell(row=v+u, column=20).value = traff_period[v]
        ws.cell(row=v+u, column=21).value = costs_per_key[v]
        ws.cell(row=v+u, column=22).value = count_conv[v]
        ws.cell(row=v+u, column=23).value = conv_cost[v]
        ws.cell(row=v+u, column=24).value = cost_click[v]
        ws.cell(row=v+u, column=28).value = full_costs_per_key[v]
        ws.cell(row=v+u, column=29).value = full_conv_per_key[v]
        if full_conv_per_key[v] > 0:
            ws.cell(row=v+u, column=30).value = full_costs_per_key[v] / full_conv_per_key[v]


    #считаем расходы и конверсии у фиксированной ставки
    cost_conv_period_fix = money_fix / all_count_conv
    all_money_fix += money_fix

    #считаем расходы и конверсии у оптимизатора
    cost_conv_period = money / all_count_conv
    all_money += money

    #ввывод итоговых цифр за период
    ws.cell(row=l, column=14).value = money_fix
    ws.cell(row=l, column=21).value = money
    ws.cell(row=l, column=15).value = all_count_conv_fix
    ws.cell(row=l, column=22).value = all_count_conv
    ws.cell(row=l, column=16).value = cost_conv_period_fix
    ws.cell(row=l, column=23).value = cost_conv_period

    U = 0
    full_money += money
    full_conv += all_count_conv
    
    full_money_fix += money_fix
    full_conv_fix += all_count_conv_fix
    #выставляем ставки
    for i in range(1, 50):
        if full_conv_per_key[i] > 0:
            if full_costs_per_key[i] / full_conv_per_key[i] < click_value:
                cost_click[i] = cost_click[i] + cost_click[i]*0.2
            else:
=======
>>>>>>> parent of d698773... Работающий оптимизатор v1
                cost_click[i] = cost_click[i] - cost_click[i]*0.2
                
            if full_costs_per_key[i] / full_conv_per_key[i] > click_value*2:
                cost_click[i] = 0
        else:
            if full_costs_per_key[i] > 1700:
                cost_click[i] = 0

    u += 52
    l += 52           


#        U = traff_period[i] * ver_conv[i]
#
#        if U * cpc[i] < click_value:
#            cost_click[i] = cost_click[i] + cost_click[i]*0.1
#        else:
#            cost_click[i] = cost_click[i] - cost_click[i]*0.1
            
            
<<<<<<< HEAD
>>>>>>> parent of d698773... Работающий оптимизатор v1
=======
>>>>>>> parent of d698773... Работающий оптимизатор v1

ws.cell(row=1, column=32).value = full_money
ws.cell(row=1, column=33).value = full_conv
ws.cell(row=1, column=34).value = full_money / full_conv

ws.cell(row=2, column=32).value = full_money_fix
ws.cell(row=2, column=33).value = full_conv_fix
ws.cell(row=2, column=34).value = full_money_fix / full_conv_fix

# Save the file
wb.save("sample1.xlsx")

print ("Done")
